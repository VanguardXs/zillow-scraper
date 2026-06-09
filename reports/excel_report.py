import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from sqlalchemy import func
from database.db import SessionLocal
from database.models import Listing
from datetime import datetime

def generate_report():
    db = SessionLocal()
    wb = Workbook()

    # ===== SHEET 1: All Listings =====
    ws1 = wb.active
    ws1.title = "All Listings"

    headers = [
        "ID", "Address", "City", "State", "Zipcode",
        "Price ($)", "Bedrooms", "Bathrooms", "Area (sqft)",
        "Home Type", "Status", "Zestimate ($)", "URL", "Created At"
    ]

    # Header styles
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data
    listings = db.query(Listing).all()
    row_fill_1 = PatternFill("solid", fgColor="D6E4F0")
    row_fill_2 = PatternFill("solid", fgColor="FFFFFF")

    for row, listing in enumerate(listings, 2):
        fill = row_fill_1 if row % 2 == 0 else row_fill_2
        data = [
            listing.id, listing.address, listing.city, listing.state,
            listing.zipcode, listing.price, listing.bedrooms, listing.bathrooms,
            listing.living_area, listing.home_type, listing.home_status,
            listing.zestimate, listing.url,
            listing.created_at.strftime("%Y-%m-%d") if listing.created_at else ""
        ]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [6, 35, 18, 8, 12, 15, 10, 12, 14, 15, 12, 15, 50, 14]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width

    # ===== SHEET 2: Stats by City =====
    ws2 = wb.create_sheet("Stats by City")

    stat_headers = ["City", "State", "Count", "Avg Price ($)", "Min Price ($)", "Max Price ($)", "Avg Area (sqft)"]

    for col, header in enumerate(stat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    stats = db.query(
        Listing.city,
        Listing.state,
        func.count(Listing.id).label("count"),
        func.avg(Listing.price).label("avg_price"),
        func.min(Listing.price).label("min_price"),
        func.max(Listing.price).label("max_price"),
        func.avg(Listing.living_area).label("avg_area")
    ).group_by(Listing.city, Listing.state).all()

    for row, s in enumerate(stats, 2):
        fill = row_fill_1 if row % 2 == 0 else row_fill_2
        data = [
            s.city, s.state, s.count,
            round(s.avg_price or 0, 2),
            s.min_price, s.max_price,
            round(s.avg_area or 0, 2)
        ]
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col, width in enumerate([18, 8, 8, 16, 16, 16, 16], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = width

    # ===== SHEET 3: Chart =====
    ws3 = wb.create_sheet("Chart")

    ws3.cell(row=1, column=1, value="City").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Avg Price ($)").font = Font(bold=True)

    for row, s in enumerate(stats, 2):
        ws3.cell(row=row, column=1, value=s.city)
        ws3.cell(row=row, column=2, value=round(s.avg_price or 0, 2))

    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Price by City"
    chart.y_axis.title = "Price ($)"
    chart.x_axis.title = "City"
    chart.width = 25
    chart.height = 15

    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(stats) + 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(stats) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws3.add_chart(chart, "D2")

    # Save file
    os.makedirs("output", exist_ok=True)
    filename = f"output/zillow_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    db.close()
    print(f"✅ Report saved: {filename}")
    return filename

if __name__ == "__main__":
    generate_report()